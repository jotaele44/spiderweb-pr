"""
build_geography_catalog.py — Filter the U.S. Census Bureau "Geography Collections"
spreadsheet down to the PR-applicable, workbench-relevant subset and write a
small CSV that doubles as a planning checklist for future TIGER ingestion work.

The .xlsx itself is NOT committed; this script captures only the filtered slice
plus a stable mapping table for which summary levels we've already wired into
the pipeline.

Usage:
    python3 server/ingestion/build_geography_catalog.py \\
        --src ~/Downloads/list-of-available-collections-of-geographies.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).parent.parent.parent
DEFAULT_OUT = REPO_ROOT / "data" / "census" / "geography_collections_pr.csv"

# Summary-level child IDs that are state-internal and apply to PR (PR is treated
# as a state in Census geography, STATEFP=72).
PR_APPLICABLE = {
    "040", "050", "060", "100", "140", "150", "160", "170",
    "310", "330", "400", "420", "500", "795", "860", "861",
    "871", "950", "960", "970",
}

# Of the above, this is the "core" subset that maps onto the workbench's
# finance / spatial / anomaly remit. Other rows are kept in the CSV but flagged
# as not-yet-wired-in-repo.
#
# SL 871 is excluded: it's "ZCTA5 within State" — a state-nested rollup query
# over the same data SL 860 covers nationally. Once 860 is wired and filtered
# to PR (STATEFP=72-ish prefix), 871 adds nothing operationally.
# PR's 8 planning regions are NOT a Census-tabulated geography; they come from
# Junta de Planificación and are tracked as a B2 follow-on, not in this CSV.
CORE_SL = {
    "040", "050", "060", "140", "150", "160", "310", "330",
    "795", "860", "950", "960", "970",
}

# Layers currently ingested by server/ingestion/ingest_tiger_pr.py.
WIRED_TODAY = {"040", "050", "060", "140", "150", "160", "860"}

# Short human label for each SL — same legend used in the reference doc.
SL_LABEL = {
    "040": "State / Statistically Equivalent (PR as a whole)",
    "050": "County (PR: Municipio)",
    "060": "County Subdivision (PR: Barrio-pueblo / Barrio)",
    "100": "Block",
    "140": "Census Tract",
    "150": "Block Group",
    "160": "Place (Incorporated / CDP)",
    "170": "Consolidated City",
    "310": "Metro/Micro Statistical Area (CBSA)",
    "330": "Combined Statistical Area",
    "400": "Urban Area",
    "420": "Urban-Rural breakdown",
    "500": "Congressional District",
    "795": "Public Use Microdata Area (PUMA)",
    "860": "ZCTA (ZIP Code Tabulation Area)",
    "861": "ZCTA-5 nested",
    "871": "ZCTA5 within State (state-nested variant of 860)",
    "950": "School District (Elementary)",
    "960": "School District (Secondary)",
    "970": "School District (Unified)",
}

NOTES_BY_SL = {
    "100": "Useful but ~1M features for PR — out of scope until a consumer needs it",
}


def _norm_sl(value) -> str:
    """Normalize a summary-level value (some xlsx cells are int, some str)."""
    if value is None:
        return ""
    return str(value).strip().zfill(3) if str(value).strip().isdigit() else str(value).strip()


def filter_rows(src: Path) -> list[dict]:
    wb = load_workbook(src, data_only=True)
    ws = wb["Geography Collections"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # header

    out: list[dict] = []
    for r in rows_iter:
        child_sl = _norm_sl(r[0])
        child_geoid_template = str(r[1]).strip() if r[1] is not None else ""
        parent_sl = _norm_sl(r[2])
        description = str(r[3]).strip() if r[3] is not None else ""

        if child_sl not in CORE_SL:
            continue

        notes = NOTES_BY_SL.get(child_sl, "")
        out.append({
            "child_sl": child_sl,
            "child_sl_label": SL_LABEL.get(child_sl, ""),
            "child_geoid_template": child_geoid_template,
            "parent_sl": parent_sl,
            "description": description,
            "wired_in_repo": "true" if child_sl in WIRED_TODAY else "false",
            "notes": notes,
        })

    out.sort(key=lambda r: (r["child_sl"], r["parent_sl"]))
    return out


def write_csv(rows: list[dict], dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "child_sl", "child_sl_label", "child_geoid_template",
        "parent_sl", "description", "wired_in_repo", "notes",
    ]
    with dest.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--src",
        type=Path,
        default=Path.home() / "Downloads" / "list-of-available-collections-of-geographies.xlsx",
        help="Path to the Census Geography Collections xlsx",
    )
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = p.parse_args(argv)

    if not args.src.exists():
        print(f"FATAL: source not found: {args.src}", file=sys.stderr)
        return 2

    rows = filter_rows(args.src)
    write_csv(rows, args.out)
    wired = sum(1 for r in rows if r["wired_in_repo"] == "true")
    print(
        f"wrote {args.out.relative_to(REPO_ROOT)} — "
        f"{len(rows)} rollups ({wired} wired, {len(rows) - wired} gaps)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
